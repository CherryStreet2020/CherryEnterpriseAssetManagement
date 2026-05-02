#!/usr/bin/env python3
"""
=============================================================================
DEPRECATED / DO NOT RUN — Use /Admin/DataImport seed pipelines instead.
=============================================================================

WARNING: This script contains DELETE operations that will destroy data!
         Line 85: DELETE FROM "Assets" (DELETES ALL ASSETS!)

This script has been deprecated as of 2026-01-22 (Phase 1 Safety Hardening).
Use the C# seed pipelines via /Admin/DataImport for safe, idempotent seeding.

=============================================================================
"""

import os
import sys

# =============================================================================
# SAFETY GUARD: Block execution unless explicit override is provided
# =============================================================================
def check_safety_guards():
    """Abort unless explicit safety overrides are set."""
    
    # Check 1: Require explicit acknowledgment
    allow_dangerous = os.environ.get('ALLOW_DANGEROUS_SEED_SCRIPTS', '')
    if allow_dangerous != 'I_UNDERSTAND_THIS_CAN_DELETE_DATA':
        print("=" * 70)
        print("BLOCKED: This script is DEPRECATED and can DELETE DATA.")
        print("=" * 70)
        print()
        print("This script contains DELETE FROM operations:")
        print("  - DELETE FROM \"Assets\" (DELETES ALL ASSETS!)")
        print()
        print("Use /Admin/DataImport in the web UI instead (safe, idempotent).")
        print()
        print("If you REALLY need to run this script, set:")
        print("  export ALLOW_DANGEROUS_SEED_SCRIPTS=I_UNDERSTAND_THIS_CAN_DELETE_DATA")
        print("  export ASPNETCORE_ENVIRONMENT=Development")
        print("=" * 70)
        sys.exit(1)
    
    # Check 2: Require LAB/Development environment
    env = os.environ.get('ASPNETCORE_ENVIRONMENT', '')
    if env.lower() not in ('development', 'lab'):
        print("=" * 70)
        print("BLOCKED: This script can only run in Development/LAB environment.")
        print(f"Current ASPNETCORE_ENVIRONMENT: {env or '(not set)'}")
        print("=" * 70)
        sys.exit(1)
    
    print("=" * 70)
    print("WARNING: Safety guards passed. Proceeding with DESTRUCTIVE operations.")
    print("=" * 70)

# Run safety check immediately on import
check_safety_guards()

import pg8000
from openpyxl import load_workbook
from datetime import datetime

def main():
    host = os.environ.get('PGHOST')
    port = int(os.environ.get('PGPORT', 5432))
    user = os.environ.get('PGUSER')
    password = os.environ.get('PGPASSWORD')
    database = os.environ.get('PGDATABASE')
    conn = pg8000.connect(host=host, port=port, user=user, password=password, database=database)
    
    cur = conn.cursor()
    
    print("Loading Excel file...")
    wb = load_workbook('attached_assets/ABS_FA_FY25_1768664094611.xlsx', data_only=True)
    ws = wb['SUMMARY']
    
    assets = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=9, values_only=True), start=9):
        asset_num = row[0]
        if asset_num is None or not isinstance(asset_num, (int, float)):
            continue
        
        description = str(row[1] or '').strip()
        if not description or description == 'NOT IN USE - SOLD':
            continue
            
        model = str(row[2]) if row[2] else None
        serial = str(row[3]) if row[3] else None
        location = str(row[5]) if row[5] else None
        bay = str(row[6]) if row[6] else None
        fiscal_year_raw = row[7]
        if fiscal_year_raw:
            try:
                if isinstance(fiscal_year_raw, (int, float)):
                    fiscal_year = int(fiscal_year_raw)
                else:
                    fiscal_year = int(str(fiscal_year_raw).split('-')[0])
            except:
                fiscal_year = None
        else:
            fiscal_year = None
        def safe_float(val, default=0):
            if val is None:
                return default
            try:
                return float(val)
            except (ValueError, TypeError):
                return default
        
        acquisition_cost = safe_float(row[8], 0)
        accumulated_depreciation = safe_float(row[9], 0)
        book_value = safe_float(row[10], 0)
        fmv = safe_float(row[11], None)
        active_str = str(row[12] or '').strip().upper()
        active = active_str in ('A', 'ACTIVE', '')
        
        in_service_date = datetime(fiscal_year, 4, 30) if fiscal_year else datetime(2000, 1, 1)
        
        assets.append({
            'asset_number': str(int(asset_num)),
            'description': description[:200],
            'model': model[:200] if model else None,
            'serial_number': serial[:100] if serial else None,
            'location': location[:100] if location else None,
            'bay': bay[:50] if bay else None,
            'fiscal_purchase_year': fiscal_year,
            'in_service_date': in_service_date,
            'acquisition_cost': acquisition_cost,
            'accumulated_depreciation': accumulated_depreciation,
            'salvage_value': 0,
            'book_value': book_value,
            'fair_market_value': fmv,
            'active': active,
            'currency': 'CAD',
            'depreciation_method': 0,
            'useful_life_months': 120,
        })
    
    print(f"Found {len(assets)} assets to import")
    
    cur.execute('DELETE FROM "Assets"')
    print("Cleared existing assets")
    
    insert_sql = '''
        INSERT INTO "Assets" (
            "AssetNumber", "Description", "Model", "SerialNumber", 
            "Location", "Bay", "FiscalPurchaseYear", "InServiceDate",
            "AcquisitionCost", "AccumulatedDepreciation", "SalvageValue", 
            "BookValue", "FairMarketValue", "Active", "Currency", 
            "DepreciationMethod", "UsefulLifeMonths"
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    '''
    
    for a in assets:
        cur.execute(insert_sql, (
            a['asset_number'], a['description'], a['model'], a['serial_number'],
            a['location'], a['bay'], a['fiscal_purchase_year'], a['in_service_date'],
            a['acquisition_cost'], a['accumulated_depreciation'], a['salvage_value'],
            a['book_value'], a['fair_market_value'], a['active'], a['currency'],
            a['depreciation_method'], a['useful_life_months']
        ))
    
    conn.commit()
    print(f"Successfully imported {len(assets)} assets!")
    
    cur.execute('SELECT COUNT(*) FROM "Assets"')
    count = cur.fetchone()[0]
    print(f"Total assets in database: {count}")
    
    cur.close()
    conn.close()

if __name__ == '__main__':
    main()
